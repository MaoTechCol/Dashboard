from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone as datetime_timezone
import json
from pathlib import Path
import re
from typing import Any
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.database import SessionLocal
from app.core.time import ensure_utc, utc_now
from app.models import (
    AlarmEvent,
    DataCertificationRun,
    DailyMileageSnapshot,
    HowenAlarmRaw,
    ReconciliationReview,
)
from app.services.company_registry import CompanyRegistry
from app.core.settings import get_settings
from app.services.howen import HISTORICAL_ALARM_TYPE_MAP


def _text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _normalized_alarm_type(value: Any) -> str:
    return _text(value).lower()


def _as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    raw = _text(value)
    if not raw:
        return None
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            continue
    return None


def _utc_second(value: datetime | None, timezone: ZoneInfo) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone)
    return value.astimezone(datetime_timezone.utc).replace(microsecond=0)


def _as_km(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return round(float(value), 3)
    raw = _text(value).lower().replace("km", "").replace(" ", "")
    if not raw:
        return None
    if "," in raw and "." in raw:
        raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return round(float(raw), 3)
    except ValueError:
        return None


def _export_date_from_filename(path: Path) -> date | None:
    match = re.search(r"(20\d{6})", path.stem)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def _event_key(device_id: Any, category: Any, occurred_at: datetime | None) -> tuple[str, str, str] | None:
    if not device_id or not category or occurred_at is None:
        return None
    return (_text(device_id), _text(category), occurred_at.isoformat())


def _counter_difference(left: Counter, right: Counter) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key, count in (left - right).most_common():
        device_id, category, occurred_at = key
        rows.append(
            {
                "device_id": device_id,
                "category": category,
                "occurred_at": occurred_at,
                "count": count,
            }
        )
    return rows


def read_alarm_export(path: Path, *, fleet_name: str | None, timezone: ZoneInfo) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(rows)]
    positions = {header: index for index, header in enumerate(headers)}
    required = {"Device ID", "Alarm Type", "Begin Time"}
    missing = sorted(required - positions.keys())
    if missing:
        raise ValueError(f"Alarm export is missing columns: {', '.join(missing)}")

    physical_rows = 0
    dms_rows = 0
    counter: Counter[tuple[str, str, str]] = Counter()
    categories: Counter[str] = Counter()
    observed: list[datetime] = []
    fleet_values: Counter[str] = Counter()
    records_by_primary: dict[tuple[str, str, str], set[tuple[str, str, str]]] = {}
    for row in rows:
        fleet = _text(row[positions["Fleet"]]) if "Fleet" in positions else ""
        if fleet_name and fleet.lower() != fleet_name.strip().lower():
            continue
        physical_rows += 1
        if fleet:
            fleet_values[fleet] += 1
        raw_type = _normalized_alarm_type(row[positions["Alarm Type"]])
        category = HISTORICAL_ALARM_TYPE_MAP.get(raw_type)
        if not category:
            continue
        occurred_at = _utc_second(_as_datetime(row[positions["Begin Time"]]), timezone)
        key = _event_key(row[positions["Device ID"]], category, occurred_at)
        if key is None:
            continue
        dms_rows += 1
        counter[key] += 1
        candidate_keys = records_by_primary.setdefault(key, {key})
        for column_name, candidate_timezone in (
            ("End Time", timezone),
            ("Reporting time", ZoneInfo("UTC")),
            ("Reporting Time", ZoneInfo("UTC")),
            ("Reporting time", timezone),
            ("Reporting Time", timezone),
        ):
            if column_name not in positions:
                continue
            candidate_at = _utc_second(_as_datetime(row[positions[column_name]]), candidate_timezone)
            candidate_key = _event_key(row[positions["Device ID"]], category, candidate_at)
            if candidate_key is not None:
                candidate_keys.add(candidate_key)
        categories[category] += 1
        observed.append(occurred_at)
    workbook.close()
    if not observed:
        raise ValueError("Alarm export does not contain DMS rows for the selected fleet")
    return {
        "path": str(path),
        "purpose": "external_benchmark",
        "mutates_operational_data": False,
        "physical_rows": physical_rows,
        "dms_rows": dms_rows,
        "unique_dms_rows": len(counter),
        "provider_duplicates": dms_rows - len(counter),
        "range_start": min(observed),
        "range_end": max(observed) + timedelta(seconds=1),
        "counter": counter,
        "records": [
            {
                "primary_key": primary,
                "candidate_keys": (primary, *sorted(candidates - {primary})),
            }
            for primary, candidates in records_by_primary.items()
        ],
        "categories": dict(categories),
        "fleets": dict(fleet_values),
    }


def read_mileage_export(path: Path, *, fleet_name: str | None) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    names = [_text(value) for value in headers]
    positions = {name: index for index, name in enumerate(names)}
    if {"Device No.", "Begin Time", "End Time"}.issubset(positions):
        return _read_mileage_record_rows(workbook, rows, positions, path=path, fleet_name=fleet_name)
    required = {"Device ID", "Total"}
    missing = sorted(required - positions.keys())
    if missing:
        workbook.close()
        raise ValueError(f"Mileage export is missing columns: {', '.join(missing)}")
    date_columns: list[tuple[int, date]] = []
    has_monthly_column = False
    for index, value in enumerate(headers):
        parsed = _as_datetime(value)
        if parsed is None and re.fullmatch(r"20\d{2}-\d{2}", _text(value)):
            parsed = datetime.strptime(_text(value), "%Y-%m")
            has_monthly_column = True
        if parsed is not None and index > positions["Total"]:
            date_columns.append((index, parsed.date()))
    if not date_columns:
        raise ValueError("Mileage export does not contain daily columns")

    by_device: dict[str, float] = {}
    daily_by_device: dict[str, dict[str, float | None]] = {}
    for row in rows:
        fleet = _text(row[positions["Fleet Name"]]) if "Fleet Name" in positions else ""
        if fleet_name and fleet.lower() != fleet_name.strip().lower():
            continue
        device_id = _text(row[positions["Device ID"]])
        if not device_id:
            continue
        total_km = _as_km(row[positions["Total"]])
        if total_km is None:
            continue
        by_device[device_id] = round(total_km, 3)
        daily_by_device[device_id] = {}
        for index, day in date_columns:
            raw = row[index]
            value = _as_km(raw)
            daily_by_device[device_id][day.isoformat()] = value
    workbook.close()
    range_end = max(day for _, day in date_columns)
    export_date = _export_date_from_filename(path)
    if has_monthly_column and export_date is not None and export_date >= range_end:
        range_end = export_date
    return {
        "path": str(path),
        "purpose": "external_benchmark",
        "mutates_operational_data": False,
        "range_start": min(day for _, day in date_columns),
        "range_end": range_end,
        "device_count": len(by_device),
        "total_km": round(sum(by_device.values()), 3),
        "by_device": by_device,
        "daily_by_device": daily_by_device,
        "format": "monthly_summary" if has_monthly_column else "daily_summary",
    }


def _read_mileage_record_rows(
    workbook,
    rows,
    positions: dict[str, int],
    *,
    path: Path,
    fleet_name: str | None,
) -> dict[str, Any]:
    by_device: Counter[str] = Counter()
    daily_by_device: dict[str, dict[str, float | None]] = {}
    observed_days: list[date] = []
    negative_rows: list[dict[str, Any]] = []
    for row in rows:
        fleet = _text(row[positions["Fleet"]]) if "Fleet" in positions else ""
        if fleet_name and fleet.lower() != fleet_name.strip().lower():
            continue
        device_id = _text(row[positions["Device No."]])
        begin_at = _as_datetime(row[positions["Begin Time"]])
        end_at = _as_datetime(row[positions["End Time"]])
        if not device_id or (begin_at is None and end_at is None):
            continue
        start_km = _as_km(row[positions["Start mileage"]]) if "Start mileage" in positions else None
        end_km = _as_km(row[positions["End mileage"]]) if "End mileage" in positions else None
        reported_km = _as_km(row[positions["Driving distance"]]) if "Driving distance" in positions else None
        distance_km = round(end_km - start_km, 3) if start_km is not None and end_km is not None else reported_km
        if distance_km is None:
            continue
        day = (end_at or begin_at).date()
        if distance_km < 0:
            negative_rows.append(
                {
                    "device_id": device_id,
                    "day": day.isoformat(),
                    "start_km": start_km,
                    "end_km": end_km,
                    "distance_km": distance_km,
                }
            )
        observed_days.append(day)
        by_device[device_id] += distance_km
        device_days = daily_by_device.setdefault(device_id, {})
        device_days[day.isoformat()] = round(float(device_days.get(day.isoformat()) or 0.0) + distance_km, 3)
    workbook.close()
    if not observed_days:
        raise ValueError("Mileage record export does not contain rows for the selected fleet")
    return {
        "path": str(path),
        "purpose": "external_benchmark",
        "mutates_operational_data": False,
        "range_start": min(observed_days),
        "range_end": max(observed_days),
        "device_count": len(by_device),
        "total_km": round(sum(by_device.values()), 3),
        "by_device": {device_id: round(value, 3) for device_id, value in by_device.items()},
        "daily_by_device": daily_by_device,
        "format": "mileage_record_odometer_delta",
        "negative_rows": negative_rows,
    }


def _match_provider_records(
    records: list[dict[str, Any]],
    local_counter: Counter[tuple[str, str, str]],
) -> tuple[list[dict[str, Any]], Counter[tuple[str, str, str]], int]:
    remaining = local_counter.copy()
    missing: Counter[tuple[str, str, str]] = Counter()
    alternate_matches = 0
    for record in records:
        primary = record["primary_key"]
        matched = None
        for candidate in record["candidate_keys"]:
            if remaining[candidate] > 0:
                matched = candidate
                break
        if matched is None:
            missing[primary] += 1
            continue
        remaining[matched] -= 1
        if remaining[matched] <= 0:
            del remaining[matched]
        if matched != primary:
            alternate_matches += 1
    return _counter_difference(missing, Counter()), remaining, alternate_matches


def certify(
    *,
    company_slug: str,
    alarm_export: dict[str, Any] | None,
    mileage_export: dict[str, Any] | None,
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "company_slug": company_slug,
        "generated_at": utc_now().isoformat(),
        "purpose": "external_benchmark",
        "mutates_operational_data": False,
    }
    range_start: datetime | None = alarm_export["range_start"] if alarm_export else None
    range_end: datetime | None = alarm_export["range_end"] if alarm_export else None

    if alarm_export:
        with SessionLocal() as session:
            raw_rows = list(
                session.scalars(
                    select(HowenAlarmRaw).where(
                        HowenAlarmRaw.company_slug == company_slug,
                        HowenAlarmRaw.occurred_at >= range_start,
                        HowenAlarmRaw.occurred_at < range_end,
                    )
                )
            )
            analytic_rows = list(
                session.scalars(
                    select(AlarmEvent).where(
                        AlarmEvent.company_slug == company_slug,
                        AlarmEvent.occurred_at >= range_start,
                        AlarmEvent.occurred_at < range_end,
                    )
                )
            )
            reviews = list(
                session.scalars(
                    select(ReconciliationReview).where(
                        ReconciliationReview.company_slug == company_slug,
                        ReconciliationReview.observed_at >= range_start,
                        ReconciliationReview.observed_at < range_end,
                    )
                )
            )

        raw_counter: Counter[tuple[str, str, str]] = Counter()
        temporal_counter: Counter[tuple[str, str, str]] = Counter()
        for row in raw_rows:
            if row.classification_status != "classified_dms":
                continue
            key = _event_key(row.device_id, row.mapped_category, ensure_utc(row.occurred_at))
            if key is not None:
                raw_counter[key] += 1
                if row.temporal_status != "accepted":
                    temporal_counter[key] += 1
        analytic_counter: Counter[tuple[str, str, str]] = Counter()
        for row in analytic_rows:
            key = _event_key(row.device_id, row.category, ensure_utc(row.occurred_at))
            if key is not None:
                analytic_counter[key] += 1

        provider_records = alarm_export["records"]
        missing_raw, remaining_raw, alternate_raw_matches = _match_provider_records(provider_records, raw_counter)
        missing_analytic, remaining_analytic, alternate_analytic_matches = _match_provider_records(
            provider_records,
            analytic_counter,
        )
        extra_raw = _counter_difference(remaining_raw, Counter())
        extra_analytic = _counter_difference(remaining_analytic, Counter())
        status_counts = Counter(review.review_status for review in reviews)
        raw_source_counts = Counter(str(row.source or "unknown") for row in raw_rows)
        raw_unexplained = sum(item["count"] for item in missing_raw + extra_raw)
        analytic_unexplained = sum(item["count"] for item in missing_analytic + extra_analytic)
        result["alarms"] = {
            "source_file": alarm_export["path"],
            "range_start": range_start.isoformat(),
            "range_end_exclusive": range_end.isoformat(),
            "provider_physical_rows": alarm_export["physical_rows"],
            "provider_dms_rows": alarm_export["dms_rows"],
            "provider_unique_dms": alarm_export["unique_dms_rows"],
            "provider_duplicate_rows": alarm_export["provider_duplicates"],
            "local_raw_dms_rows": sum(raw_counter.values()),
            "local_unique_raw_dms": len(raw_counter),
            "local_analytic_rows": sum(analytic_counter.values()),
            "local_unique_analytic": len(analytic_counter),
            "temporal_dms_rows": sum(temporal_counter.values()),
            "raw_source_counts": dict(raw_source_counts),
            "review_status_counts": dict(status_counts),
            "provider_categories": alarm_export["categories"],
            "missing_from_local_raw": missing_raw,
            "extra_in_local_raw": extra_raw,
            "missing_from_analytic": missing_analytic,
            "extra_in_analytic": extra_analytic,
            "provider_time_normalized_raw_matches": alternate_raw_matches,
            "provider_time_normalized_analytic_matches": alternate_analytic_matches,
            "raw_unexplained_alarm_count": raw_unexplained,
            "analytic_unexplained_alarm_count": analytic_unexplained,
            "unexplained_alarm_count": raw_unexplained + analytic_unexplained,
        }

    if mileage_export:
        start_date = mileage_export["range_start"]
        end_date = mileage_export["range_end"]
        with SessionLocal() as session:
            snapshots = list(
                session.scalars(
                    select(DailyMileageSnapshot).where(
                        DailyMileageSnapshot.company_slug == company_slug,
                        DailyMileageSnapshot.snapshot_date >= start_date,
                        DailyMileageSnapshot.snapshot_date <= end_date,
                    )
                )
            )
        local_by_device: Counter[str] = Counter()
        excluded_by_device: Counter[str] = Counter()
        missing_days_by_device: Counter[str] = Counter()
        expected_days = (end_date - start_date).days + 1
        covered_days: Counter[str] = Counter()
        for snapshot in snapshots:
            if snapshot.excluded_at is not None or snapshot.km_validation_status != "valid":
                excluded_by_device[snapshot.device_id] += 1
                continue
            if snapshot.day_km is None:
                missing_days_by_device[snapshot.device_id] += 1
                continue
            local_by_device[snapshot.device_id] += float(snapshot.day_km)
            covered_days[snapshot.device_id] += 1

        comparisons: list[dict[str, Any]] = []
        provider_devices = mileage_export["by_device"]
        for device_id in sorted(set(provider_devices) | set(local_by_device)):
            provider_km = round(float(provider_devices.get(device_id, 0.0)), 3)
            local_km = round(float(local_by_device.get(device_id, 0.0)), 3)
            comparisons.append(
                {
                    "device_id": device_id,
                    "provider_km": provider_km,
                    "local_km": local_km,
                    "difference_km": round(local_km - provider_km, 3),
                    "covered_days": covered_days.get(device_id, 0),
                    "expected_days": expected_days,
                    "excluded_days": excluded_by_device.get(device_id, 0),
                    "missing_days": max(expected_days - covered_days.get(device_id, 0) - excluded_by_device.get(device_id, 0), 0),
                }
            )
        provider_total = round(float(mileage_export["total_km"]), 3)
        local_total = round(sum(local_by_device.values()), 3)
        difference_pct = round(abs(local_total - provider_total) / provider_total * 100, 4) if provider_total else None
        result["mileage"] = {
            "source_file": mileage_export["path"],
            "source_format": mileage_export["format"],
            "range_start": start_date.isoformat(),
            "range_end": end_date.isoformat(),
            "provider_device_count": mileage_export["device_count"],
            "provider_km": provider_total,
            "local_km": local_total,
            "difference_km": round(local_total - provider_total, 3),
            "difference_pct": difference_pct,
            "vehicles": comparisons,
        }

    alarm_ok = not alarm_export or (
        result["alarms"]["raw_unexplained_alarm_count"] == 0
        and result["alarms"]["analytic_unexplained_alarm_count"] == 0
    )
    km_ok = not mileage_export or (
        result["mileage"]["difference_pct"] is not None and result["mileage"]["difference_pct"] < 1.0
    )
    result["status"] = "passed" if alarm_ok and km_ok else "failed"
    result["acceptance"] = {
        "alarm_unexplained_zero": alarm_ok,
        "alarm_raw_layer_clean": not alarm_export
        or result["alarms"]["raw_unexplained_alarm_count"] == 0,
        "alarm_analytic_layer_clean": not alarm_export
        or result["alarms"]["analytic_unexplained_alarm_count"] == 0,
        "mileage_difference_below_1pct": km_ok,
    }
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Certify local DMS data against Howen Excel exports")
    parser.add_argument("--company", required=True)
    parser.add_argument("--alarm-file", type=Path)
    parser.add_argument("--mileage-file", type=Path)
    parser.add_argument("--fleet-name")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    if not args.alarm_file and not args.mileage_file:
        parser.error("At least one of --alarm-file or --mileage-file is required")

    settings = get_settings()
    registry = CompanyRegistry(
        settings.company_config_path,
        seed_path=settings.company_seed_config_path,
        session_factory=SessionLocal,
    )
    try:
        company = registry.get(args.company)
    except KeyError as exc:
        result = {
            "company_slug": args.company,
            "generated_at": utc_now().isoformat(),
            "status": "blocked",
            "reason": "company_not_registered",
            "message": str(exc),
        }
        output = args.output or Path("storage") / "certifications" / f"{args.company}-blocked.json"
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(json.dumps({"status": "blocked", "output": str(output)}, ensure_ascii=False))
        return
    timezone = ZoneInfo(company.timezone or settings.default_timezone)
    alarm_export = (
        read_alarm_export(args.alarm_file, fleet_name=args.fleet_name, timezone=timezone)
        if args.alarm_file
        else None
    )
    mileage_export = (
        read_mileage_export(args.mileage_file, fleet_name=args.fleet_name)
        if args.mileage_file
        else None
    )
    result = certify(company_slug=company.slug, alarm_export=alarm_export, mileage_export=mileage_export)
    run_id = uuid4().hex
    alarm_range_start = alarm_export["range_start"] if alarm_export else None
    alarm_range_end = alarm_export["range_end"] if alarm_export else None
    if not alarm_range_start and mileage_export:
        alarm_range_start = datetime.combine(mileage_export["range_start"], time.min, tzinfo=timezone).astimezone(ZoneInfo("UTC"))
        alarm_range_end = datetime.combine(mileage_export["range_end"] + timedelta(days=1), time.min, tzinfo=timezone).astimezone(ZoneInfo("UTC"))
    with SessionLocal() as session:
        session.add(
            DataCertificationRun(
                id=run_id,
                company_slug=company.slug,
                source_name=", ".join(
                    str(path.name) for path in (args.alarm_file, args.mileage_file) if path is not None
                ),
                range_start=alarm_range_start,
                range_end=alarm_range_end,
                status=result["status"],
                provider_alarm_count=result.get("alarms", {}).get("provider_unique_dms", 0),
                local_raw_count=result.get("alarms", {}).get("local_unique_raw_dms", 0),
                local_analytic_count=result.get("alarms", {}).get("local_unique_analytic", 0),
                unexplained_alarm_count=result.get("alarms", {}).get("unexplained_alarm_count", 0),
                provider_km=result.get("mileage", {}).get("provider_km"),
                local_km=result.get("mileage", {}).get("local_km"),
                km_difference_pct=result.get("mileage", {}).get("difference_pct"),
                result_json=json.dumps(result, ensure_ascii=True, default=str),
            )
        )
        session.commit()
    result["certification_run_id"] = run_id
    output = args.output or Path("storage") / "certifications" / f"{company.slug}-{run_id}.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "run_id": run_id, "output": str(output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
