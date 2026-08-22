from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import importlib.util
import unittest
from zoneinfo import ZoneInfo

from openpyxl import Workbook

MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts" / "certify_real_data.py"
SPEC = importlib.util.spec_from_file_location("certify_real_data", MODULE_PATH)
assert SPEC and SPEC.loader
CERTIFICATION = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(CERTIFICATION)
read_alarm_export = CERTIFICATION.read_alarm_export
read_mileage_export = CERTIFICATION.read_mileage_export


class CertificationExportTests(unittest.TestCase):
    def test_alarm_export_counts_physical_duplicates_separately(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "alarms.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Device ID", "Alarm Type", "Fleet", "Begin Time"])
            sheet.append(["device-1", "Eyes Closed", "ISMOCOL UTIJP", "2026-08-21 10:00:00"])
            sheet.append(["device-1", "Eyes Closed", "ISMOCOL UTIJP", "2026-08-21 10:00:00"])
            sheet.append(["device-1", "Ignition On", "ISMOCOL UTIJP", "2026-08-21 10:01:00"])
            workbook.save(path)

            result = read_alarm_export(
                path,
                fleet_name="ISMOCOL UTIJP",
                timezone=ZoneInfo("America/Bogota"),
            )

        self.assertEqual(result["physical_rows"], 3)
        self.assertEqual(result["dms_rows"], 2)
        self.assertEqual(result["unique_dms_rows"], 1)
        self.assertEqual(result["provider_duplicates"], 1)
        self.assertEqual(result["purpose"], "external_benchmark")
        self.assertFalse(result["mutates_operational_data"])
        key = result["records"][0]["primary_key"]
        self.assertEqual(key[2], "2026-08-21T15:00:00+00:00")

    def test_alarm_export_keeps_begin_time_as_primary_candidate(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "alarms-with-reporting-time.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Device ID", "Alarm Type", "Fleet", "Begin Time", "Reporting time"])
            sheet.append(
                [
                    "device-1",
                    "Eyes Closed",
                    "ISMOCOL UTIJP",
                    "2026-08-21 10:00:00",
                    "2026-08-21 09:59:00",
                ]
            )
            workbook.save(path)

            result = read_alarm_export(
                path,
                fleet_name="ISMOCOL UTIJP",
                timezone=ZoneInfo("America/Bogota"),
            )

        record = result["records"][0]
        self.assertEqual(record["candidate_keys"][0], record["primary_key"])
        self.assertGreater(len(record["candidate_keys"]), 1)

    def test_mileage_export_preserves_explicit_zero_and_missing_cells(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "mileage.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Device ID", "Device Name", "Fleet Name", "Total", "2026-08-20", "2026-08-21"])
            sheet.append(["device-1", "ABC123", "ISMOCOL UTIJP", "15.5 km", "15.5 km", "0 km"])
            sheet.append(["device-2", "DEF456", "ISMOCOL UTIJP", "7.0 km", "7.0 km", None])
            workbook.save(path)

            result = read_mileage_export(path, fleet_name="ISMOCOL UTIJP")

        self.assertEqual(result["device_count"], 2)
        self.assertEqual(result["total_km"], 22.5)
        self.assertEqual(result["daily_by_device"]["device-1"]["2026-08-21"], 0.0)
        self.assertIsNone(result["daily_by_device"]["device-2"]["2026-08-21"])

    def test_mileage_record_uses_odometer_delta_instead_of_bad_driving_distance(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "mileage-record.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "Device No.",
                    "Fleet",
                    "Driving distance",
                    "Begin Time",
                    "Start mileage",
                    "End Time",
                    "End mileage",
                ]
            )
            sheet.append(
                [
                    "device-1",
                    "ISMOCOL UTIJP",
                    "-977.07 km",
                    "2026-08-21 00:00:00",
                    "1000.00 km",
                    "2026-08-21 23:59:00",
                    "3076.82 km",
                ]
            )
            workbook.save(path)

            result = read_mileage_export(path, fleet_name="ISMOCOL UTIJP")

        self.assertEqual(result["format"], "mileage_record_odometer_delta")
        self.assertEqual(result["device_count"], 1)
        self.assertEqual(result["total_km"], 2076.82)
        self.assertEqual(result["negative_rows"], [])

    def test_monthly_summary_uses_export_date_as_period_end(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "mileage_Statistic_20260817203742.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Device ID", "Device Name", "Fleet Name", "Total", "2026-08"])
            sheet.append(["device-1", "ABC123", "ISMOCOL UTIJP", "2248.82 km", "2248.82 km"])
            workbook.save(path)

            result = read_mileage_export(path, fleet_name="ISMOCOL UTIJP")

        self.assertEqual(result["format"], "monthly_summary")
        self.assertEqual(result["range_start"].isoformat(), "2026-08-01")
        self.assertEqual(result["range_end"].isoformat(), "2026-08-17")
        self.assertEqual(result["total_km"], 2248.82)


if __name__ == "__main__":
    unittest.main()
